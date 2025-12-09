import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

import warnings
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=FutureWarning)

import sys
import re
import numpy as np
import cv2
from keras.models import load_model
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


# ---------- 參數 ----------
# ---------- 參數 ----------
IMAGE_SIZE = (128, 128)
MSE_THRESHOLD = 0.0028  # 缺陷判斷閾值
SCRIPT_DIR = Path(__file__).parent
MODEL_PATH = SCRIPT_DIR / 'led_autoencoder(1).h5'
DIFF_FOLDER_NAME = 'diff'
EXCEL_FILENAME = 'result.xlsx'

# ---------- 工具函數 ----------
def extract_grid_size(folder_name):
    match = re.search(r'(\d+)x(\d+)$', folder_name)
    if match:
        return int(match.group(1)), int(match.group(2))
    else:
        raise ValueError("資料夾名稱必須包含矩陣格式，例如 '2025-04-30-20-10-00_5x5'")

def load_and_preprocess_image(image_path):
    img = cv2.imread(str(image_path))
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img = cv2.resize(img, IMAGE_SIZE)
    img = img.astype('float32') / 255.0
    return img

def generate_diff_image(original, reconstructed):
    diff = np.abs(original - reconstructed)
    diff = (diff * 255).astype(np.uint8)
    return diff

# ---------- 主程式 ----------
def main():
    if len(sys.argv) != 3:
        print("使用方法: python analyze.py <輸入資料夾> <輸出資料夾>")
        return

    input_folder = Path(sys.argv[1])
    output_folder = Path(sys.argv[2])
    output_folder.mkdir(parents=True, exist_ok=True)
    diff_folder = output_folder / DIFF_FOLDER_NAME
    diff_folder.mkdir(exist_ok=True)

    folder_name = input_folder.name
    rows, cols = extract_grid_size(folder_name)

    print(f"正在讀取模型檔案: {MODEL_PATH.resolve()}")
    model = load_model(MODEL_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "分析結果"

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    thin_border = Border(  # ✅ 加這段定義邊框
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

    total = 0
    defect_count = 0

    for file in os.listdir(input_folder):

        if not file.lower().endswith(('.jpg', '.png')):
            continue

        match = re.match(r'r(\d+)_c(\d+)', file)
        if not match:
            continue
        r, c = int(match.group(1)), int(match.group(2))

        img_path = input_folder / file
        original = load_and_preprocess_image(img_path)
        input_img = np.expand_dims(original, axis=0)
        reconstructed = model.predict(input_img)[0]

        mse = np.mean((original - reconstructed) ** 2)
        is_defect = mse > MSE_THRESHOLD

        total += 1
        cell = ws.cell(row=r+2, column=c+1)       # 取得該格
        cell.value = f"r{r}_c{c}"                 # ✅ 顯示座標文字
        cell.border = thin_border                 # ✅ 加邊框
        if is_defect:
            defect_count += 1
            ws.cell(row=r+2, column=c+1).fill = red_fill

            diff_img = generate_diff_image(original, reconstructed)
            diff_name = f'diff_r{r}_c{c}.png'
            cv2.imwrite(str(diff_folder / diff_name), cv2.cvtColor(diff_img, cv2.COLOR_RGB2BGR))

    ok_count = total - defect_count

    # 在表頭寫入統計資訊
    ws.cell(row=1, column=1, value=f"良品數量: {ok_count}")
    ws.cell(row=1, column=2, value=f"缺陷數量: {defect_count}")

    # 自動調整欄寬
    for col in range(1, cols+1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(output_folder / EXCEL_FILENAME)
    print(f"完成分析，共處理 {total} 張圖片，其中 {defect_count} 張為缺陷。")

if __name__ == '__main__':
    main()
